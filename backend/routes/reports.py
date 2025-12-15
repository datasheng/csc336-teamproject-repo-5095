# backend/routes/reports.py
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
from database.queries import get_revenue_report, get_revenue_details
from database.connection import get_db_connection
import openpyxl
from io import BytesIO
from datetime import datetime

router = APIRouter(prefix="/api/reports", tags=["Reports"])

COMMISSION_RATE = 0.15


def _to_float(x: Any, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        return float(x)
    except Exception:
        return default


def _to_int(x: Any, default: int = 0) -> int:
    try:
        if x is None:
            return default
        return int(x)
    except Exception:
        return default


@router.get("/revenue")
def get_revenue_data():
    """
    Get revenue data for all restaurants
    Returns aggregated statistics per restaurant (from database)
    """
    try:
        report = get_revenue_report()

        print(f"🔍 Revenue report type: {type(report)}")
        print(f"🔍 Revenue report value: {report}")

        if report is None:
            print("⚠️ get_revenue_report() returned None")
            return []

        if not isinstance(report, list):
            print(f"⚠️ Expected list, got {type(report)}")
            return []

        # Just return the data as-is from the database query
        # The query already includes PLATFORM_COMMISSION, SERVICE_FEES, DELIVERY_PROFIT
        return report

    except Exception as e:
        print(f"❌ Error in get_revenue_data: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate revenue report: {str(e)}")

@router.get("/revenue/details")
def get_detailed_revenue():
    """
    Get detailed revenue data (order-by-order breakdown)
    Returns individual order profit data from INVESTOR_PROFIT_VIEW
    """
    try:
        details = get_revenue_details()

        if details is None:
            raise HTTPException(status_code=500, detail="Failed to connect to database")

        return {
            "success": True,
            "message": "Revenue details retrieved",
            "data": details,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get details: {str(e)}")


@router.get("/revenue/excel")
def download_revenue_excel():
    try:
        report_data = get_revenue_report()

        if not report_data:
            raise HTTPException(
                status_code=404,
                detail="No revenue data available"
            )

        COMMISSION_RATE = 0.15
        SERVICE_FEE_PER_ORDER = 2.99
        DELIVERY_COMMISSION_PER_ORDER = 0.60

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Report"

        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Platform Revenue Report - {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = openpyxl.styles.Font(size=14, bold=True)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Headers
        ws.append([])  # row 2 empty

        headers = [
            "Restaurant",
            "Total Orders",
            "Total Revenue (Gross)",
            "Avg Order Value",
            "Unique Customers",
            "Platform Commission (15%)",
            "Net to Restaurant",
            "Platform Revenue Add-ons (Service + Delivery Comm.)",
        ]
        ws.append(headers)

        for cell in ws[3]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        total_orders = 0
        total_gross = 0.0
        total_commission = 0.0
        total_net_to_restaurant = 0.0

        # Rows
        for row in report_data:
            orders = int(row.get("TOTAL_ORDERS") or 0)
            gross = float(row.get("TOTAL_REVENUE") or 0.0)
            avg = float(row.get("AVG_ORDER_VALUE") or 0.0)
            uniq = int(row.get("UNIQUE_CUSTOMERS") or 0)

            commission = row.get("PLATFORM_COMMISSION")
            if commission is None:
                commission = gross * COMMISSION_RATE
            commission = float(commission or 0.0)

            net_to_restaurant = row.get("NET_RESTAURANT_REVENUE")
            if net_to_restaurant is None:
                net_to_restaurant = gross - commission
            net_to_restaurant = float(net_to_restaurant or 0.0)

            add_ons = (orders * SERVICE_FEE_PER_ORDER) + (orders * DELIVERY_COMMISSION_PER_ORDER)

            ws.append([
                row.get("RESTAURANT_NAME"),
                orders,
                gross,
                avg,
                uniq,
                commission,
                net_to_restaurant,
                add_ons,
            ])

            total_orders += orders
            total_gross += gross
            total_commission += commission
            total_net_to_restaurant += net_to_restaurant

        # Summary section (platform earnings)
        service_fee_total = total_orders * SERVICE_FEE_PER_ORDER
        delivery_commission_total = total_orders * DELIVERY_COMMISSION_PER_ORDER
        total_platform_revenue = total_commission + service_fee_total + delivery_commission_total

        ws.append([])  # empty row
        summary_start = ws.max_row + 1

        ws[f"A{summary_start}"] = "PLATFORM EARNINGS SUMMARY"
        ws[f"A{summary_start}"].font = openpyxl.styles.Font(size=12, bold=True)

        ws[f"A{summary_start+1}"] = "Total Orders:"
        ws[f"B{summary_start+1}"] = total_orders

        ws[f"A{summary_start+2}"] = "Restaurant Commissions (15%):"
        ws[f"B{summary_start+2}"] = total_commission
        ws[f"B{summary_start+2}"].number_format = '$#,##0.00'

        ws[f"A{summary_start+3}"] = f"Service Fees (${SERVICE_FEE_PER_ORDER:.2f} / order):"
        ws[f"B{summary_start+3}"] = service_fee_total
        ws[f"B{summary_start+3}"].number_format = '$#,##0.00'

        ws[f"A{summary_start+4}"] = f"Delivery Commission (${DELIVERY_COMMISSION_PER_ORDER:.2f} / order):"
        ws[f"B{summary_start+4}"] = delivery_commission_total
        ws[f"B{summary_start+4}"].number_format = '$#,##0.00'

        ws[f"A{summary_start+5}"] = "TOTAL PLATFORM REVENUE:"
        ws[f"B{summary_start+5}"] = total_platform_revenue
        ws[f"B{summary_start+5}"].number_format = '$#,##0.00'
        ws[f"B{summary_start+5}"].font = openpyxl.styles.Font(bold=True, size=12)
        ws[f"B{summary_start+5}"].fill = openpyxl.styles.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Totals row at bottom of table (like your screenshot)
        # Table starts at row 3 header, data starts row 4
        # totals row will be row (3 + len(report_data) + 1)
        totals_row = 4 + len(report_data)

        ws[f"A{totals_row}"] = "TOTAL"
        ws[f"B{totals_row}"] = total_orders
        ws[f"C{totals_row}"] = total_gross
        ws[f"F{totals_row}"] = total_commission
        ws[f"G{totals_row}"] = total_net_to_restaurant
        ws[f"H{totals_row}"] = service_fee_total + delivery_commission_total

        for cell in ws[totals_row]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

        # Currency formats
        for r in range(4, totals_row + 1):
            ws[f"C{r}"].number_format = '$#,##0.00'
            ws[f"D{r}"].number_format = '$#,##0.00'
            ws[f"F{r}"].number_format = '$#,##0.00'
            ws[f"G{r}"].number_format = '$#,##0.00'
            ws[f"H{r}"].number_format = '$#,##0.00'

        # Auto-width columns A..H
        for col_idx in range(1, 9):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = max(12, int((max_length + 2) * 1.2))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"platform_revenue_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Excel generation failed: {str(e)}"
        )

@router.get("/restaurant/{restaurant_id}/excel")
def download_restaurant_revenue_excel(restaurant_id: int):
    """
    Download revenue report for a specific restaurant
    For restaurant owners to see their earnings
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT r.RESTAURANT_NAME, u.USER_NAME as OWNER_NAME
            FROM RESTAURANT r
            JOIN USERS u ON r.OWNER_ID = u.USER_ID
            WHERE r.RESTAURANT_ID = %s
        """,
            (restaurant_id,),
        )
        restaurant_info = cursor.fetchone()

        if not restaurant_info:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Restaurant not found")

        cursor.execute(
            """
            SELECT 
                o.ORDER_ID,
                o.ORDER_DATE,
                u.USER_NAME as CUSTOMER_NAME,
                o.TOTAL_AMOUNT as GROSS_REVENUE,
                (o.TOTAL_AMOUNT * 0.15) as PLATFORM_COMMISSION,
                (o.TOTAL_AMOUNT * 0.85) as NET_REVENUE
            FROM ORDERS o
            JOIN USERS u ON o.USER_ID = u.USER_ID 
            WHERE o.RESTAURANT_ID = %s
            ORDER BY o.ORDER_DATE DESC
        """,
            (restaurant_id,),
        )
        orders = cursor.fetchall()

        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Restaurant Revenue"

        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"REVENUE REPORT - {restaurant_info['RESTAURANT_NAME']}"
        title_cell.font = openpyxl.styles.Font(size=16, bold=True)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        ws["A3"] = "Restaurant:"
        ws["B3"] = restaurant_info["RESTAURANT_NAME"]
        ws["B3"].font = openpyxl.styles.Font(bold=True)

        ws["A4"] = "Owner:"
        ws["B4"] = restaurant_info["OWNER_NAME"]

        ws["A5"] = "Report Date:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        total_orders = len(orders)
        total_gross = sum(float(o["GROSS_REVENUE"]) for o in orders)
        total_commission = sum(float(o["PLATFORM_COMMISSION"]) for o in orders)
        total_net = sum(float(o["NET_REVENUE"]) for o in orders)

        ws["A7"] = "REVENUE SUMMARY"
        ws["A7"].font = openpyxl.styles.Font(size=12, bold=True)

        ws["A8"] = "Total Orders:"
        ws["B8"] = total_orders

        ws["A9"] = "Gross Revenue (Customer Payments):"
        ws["B9"] = total_gross
        ws["B9"].number_format = "$#,##0.00"

        ws["A10"] = "Platform Commission (15%):"
        ws["B10"] = -total_commission
        ws["B10"].number_format = "$#,##0.00"
        ws["B10"].font = openpyxl.styles.Font(color="FF0000")

        ws["A11"] = "NET REVENUE (You Keep):"
        ws["B11"] = total_net
        ws["B11"].number_format = "$#,##0.00"
        ws["B11"].font = openpyxl.styles.Font(bold=True, size=12)
        ws["B11"].fill = openpyxl.styles.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        ws["A13"] = "ORDER HISTORY"
        ws["A13"].font = openpyxl.styles.Font(size=12, bold=True)

        headers = ["Order ID", "Date", "Customer", "Gross Revenue", "Commission (15%)", "Net Revenue"]
        ws.append([])
        ws.append(headers)

        for cell in ws[15]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        for order in orders:
            ws.append(
                [
                    order["ORDER_ID"],
                    order["ORDER_DATE"].strftime("%Y-%m-%d"),
                    order["CUSTOMER_NAME"],
                    float(order["GROSS_REVENUE"]),
                    -float(order["PLATFORM_COMMISSION"]),
                    float(order["NET_REVENUE"]),
                ]
            )

        for row in ws.iter_rows(min_row=16, max_row=ws.max_row, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = "$#,##0.00"

        for col_idx in range(1, 7):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
            ws.column_dimensions[column_letter].width = max(12, int((max_length + 2) * 1.2))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{restaurant_info['RESTAURANT_NAME'].replace(' ', '_')}_Revenue_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

@router.get("/debug")
def debug_orders():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT STATUS, COUNT(*) as count FROM ORDERS GROUP BY STATUS")
    statuses = cursor.fetchall()
    
    cursor.execute("SELECT COUNT(*) as count FROM INVESTOR_PROFIT_VIEW")
    view_count = cursor.fetchone()['count']
    
    cursor.execute("SELECT SUM(TOTAL_PLATFORM_PROFIT) as total FROM INVESTOR_PROFIT_VIEW")
    view_profit = cursor.fetchone()['total']
    
    cursor.close()
    conn.close()
    
    return {
        "order_statuses": statuses,
        "investor_view_count": view_count,
        "investor_view_profit": float(view_profit) if view_profit else 0
    }


@router.get("/debug/breakdown")
def debug_breakdown():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get totals from ORDERS table
    cursor.execute("""
        SELECT 
            COUNT(*) as total_orders,
            SUM(TOTAL_AMOUNT) as gross_revenue,
            SUM(PLATFORM_COMMISSION) as total_commission,
            SUM(SERVICE_FEE) as total_service_fees,
            SUM(PLATFORM_PROFIT_ORDER) as total_order_profit
        FROM ORDERS
        WHERE STATUS = 'DELIVERED'
    """)
    orders_totals = cursor.fetchone()
    
    # Get delivery totals
    cursor.execute("""
        SELECT 
            COUNT(*) as total_deliveries,
            SUM(DELIVERY_FEE_TOTAL) as total_delivery_fees,
            SUM(DELIVERY_PLATFORM_CUT) as total_delivery_profit
        FROM DELIVERIES
    """)
    delivery_totals = cursor.fetchone()
    
    # Get investor view total
    cursor.execute("SELECT SUM(TOTAL_PLATFORM_PROFIT) as view_profit FROM INVESTOR_PROFIT_VIEW")
    view_profit = cursor.fetchone()['view_profit']
    
    cursor.close()
    conn.close()
    
    # Calculate what frontend shows
    orders = float(orders_totals['total_orders'] or 0)
    commission = float(orders_totals['total_commission'] or 0)
    service_fees_from_db = float(orders_totals['total_service_fees'] or 0)
    delivery_profit = float(delivery_totals['total_delivery_profit'] or 0)
    
    # Frontend calculation
    frontend_service_fees = orders * 2.99
    frontend_delivery_commission = orders * 0.60
    frontend_total = commission + frontend_service_fees + frontend_delivery_commission
    
    return {
        "from_database": {
            "orders": orders_totals,
            "deliveries": delivery_totals,
            "investor_view_profit": float(view_profit) if view_profit else 0
        },
        "frontend_calculation": {
            "commission_from_db": commission,
            "service_fees_calculated": frontend_service_fees,
            "delivery_commission_calculated": frontend_delivery_commission,
            "total_platform_revenue": frontend_total
        },
        "comparison": {
            "frontend_total": frontend_total,
            "tableau_total": float(view_profit) if view_profit else 0,
            "difference": frontend_total - (float(view_profit) if view_profit else 0)
        }
    }

@router.get("/debug/delivery-count")
def debug_delivery_count():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT COUNT(*) as order_count FROM ORDERS WHERE STATUS = 'DELIVERED'")
    orders = cursor.fetchone()['order_count']
    
    cursor.execute("SELECT COUNT(*) as delivery_count FROM DELIVERIES")
    deliveries = cursor.fetchone()['delivery_count']
    
    cursor.execute("SELECT COUNT(*) as matched FROM ORDERS o JOIN DELIVERIES d ON o.ORDER_ID = d.ORDER_ID WHERE o.STATUS = 'DELIVERED'")
    matched = cursor.fetchone()['matched']
    
    cursor.close()
    conn.close()
    
    return {
        "delivered_orders": orders,
        "delivery_records": deliveries,
        "matched_order_delivery_pairs": matched,
        "unmatched": orders - matched
    }

@router.get("/debug/deliveries")
def debug_deliveries():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get distinct delivery platform cut values and their counts
    cursor.execute("""
        SELECT 
            DELIVERY_PLATFORM_CUT,
            COUNT(*) as count,
            SUM(DELIVERY_PLATFORM_CUT) as subtotal
        FROM DELIVERIES
        GROUP BY DELIVERY_PLATFORM_CUT
        ORDER BY DELIVERY_PLATFORM_CUT
    """)
    breakdown = cursor.fetchall()
    
    # Get total
    cursor.execute("SELECT SUM(DELIVERY_PLATFORM_CUT) as total FROM DELIVERIES")
    total = cursor.fetchone()['total']
    
    cursor.close()
    conn.close()
    
    return {
        "delivery_cut_breakdown": breakdown,
        "total_delivery_profit": float(total) if total else 0
    }

@router.get("/debug/investor-view")
def debug_investor_view():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT 
            SUM(PLATFORM_COMMISSION) as total_commission,
            SUM(SERVICE_FEE) as total_service_fee,
            SUM(DELIVERY_PLATFORM_CUT) as total_delivery_cut,
            SUM(TOTAL_PLATFORM_PROFIT) as total_profit
        FROM INVESTOR_PROFIT_VIEW
    """)
    result = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return {
        "commission": float(result['total_commission'] or 0),
        "service_fee": float(result['total_service_fee'] or 0),
        "delivery_cut": float(result['total_delivery_cut'] or 0),
        "total_profit": float(result['total_profit'] or 0),
        "calculated_sum": float(result['total_commission'] or 0) + float(result['total_service_fee'] or 0) + float(result['total_delivery_cut'] or 0)
    }